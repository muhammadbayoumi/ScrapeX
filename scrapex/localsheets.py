"""Local .xlsx sink — the offline twin of the Google sink (ENGINEERING.md P1).

Produces a workbook with one tab per source, mirroring the Drive layout exactly,
using the SAME export_source_table data (via publish.publish_source). No Google,
no network. openpyxl is imported lazily so the rest of ScrapeX needs no xlsx dep.
"""
from __future__ import annotations

import os
from collections.abc import Iterator
from contextlib import contextmanager
from pathlib import Path
from types import ModuleType
from typing import TYPE_CHECKING

if TYPE_CHECKING:                    # openpyxl is an optional extra: types only
    from openpyxl.workbook import Workbook

DEFAULT_EXPORT_DIR = Path(os.environ.get("SCRAPEX_EXPORT_DIR", str(Path.home() / "ScrapeX")))

# Excel worksheet titles: max 31 chars, and these characters are forbidden.
_BAD_TITLE_CHARS = set(r"[]:*?/\\")


def _safe_title(tab: str) -> str:
    cleaned = "".join("_" if ch in _BAD_TITLE_CHARS else ch for ch in tab)
    return cleaned[:31] or "Sheet"


def workbook_bytes(tabs: list[tuple[str, list[str], list[list]]]) -> bytes:
    """The same workbook LocalSink writes to disk, in memory, for a download.

    The Data page's Excel button used to call Tabulator's client-side xlsx
    writer, which needs SheetJS on `window` — never vendored here, so the
    button logged a console error and produced no file at all. Worse, even
    working it could only ever have exported the grid: the details, the history
    and the provenance are not in the browser. Building the same tabs the CLI
    and the Google push build, on the server, makes one download the whole
    record (P1 DRY: publish.workbook_tables decides WHAT, this decides only how
    to write it).

    Titles are truncated to Excel's 31 characters, so two long tabs can collide
    once shortened; the suffix keeps them distinct rather than letting one
    silently replace the other.
    """
    try:
        from openpyxl import Workbook
    except ImportError as exc:  # pragma: no cover - environment guard
        raise RuntimeError("the Excel export needs: pip install -e .[local]") from exc
    from io import BytesIO

    book = Workbook()
    book.remove(book.active)             # openpyxl's default empty sheet
    used: set[str] = set()
    for tab, header, rows in tabs:
        title = _safe_title(tab)
        if title in used:
            title = f"{title[:28]}_{len(used)}"
        used.add(title)
        sheet = book.create_sheet(title)
        sheet.append(list(header))
        for row in rows:
            sheet.append(list(row))
        sheet.freeze_panes = "A2"        # the header stays put while scrolling
    stream = BytesIO()
    book.save(stream)
    return stream.getvalue()


def _openpyxl() -> ModuleType:
    """openpyxl, imported only when a workbook is actually read or written, so
    the rest of ScrapeX keeps running with no xlsx dependency installed."""
    try:
        import openpyxl
    except ImportError as exc:  # pragma: no cover - environment guard
        raise RuntimeError("local export needs: pip install -e .[local]") from exc
    return openpyxl


def _write_sheet(book: Workbook, tab: str, header: list[str], rows: list[list],
                 *, fresh: bool) -> None:
    """Put one tab into an OPEN workbook. `fresh` = the book was created here,
    so openpyxl's default empty sheet is ours to drop rather than the owner's."""
    title = _safe_title(tab)
    if title in book.sheetnames:
        del book[title]              # replace the tab (idempotent, like the Google sink)
    sheet = book.create_sheet(title)
    sheet.append(list(header))
    for row in rows:
        sheet.append(list(row))
    if fresh and "Sheet" in book.sheetnames and "Sheet" != title:
        del book["Sheet"]            # drop openpyxl's default empty sheet


def _save(book: Workbook, path: Path) -> None:
    """Write the workbook out in a step that either lands whole or not at all.

    A whole export is ONE save now (see LocalSink.batch), so a crash partway
    through writing in place would take every previously exported tab with it —
    where the per-tab save this replaced could at worst stop between tabs.
    Same write-then-rename move the databases use (databases/registry.py).
    """
    incoming = path.with_suffix(path.suffix + ".incoming")
    try:
        book.save(incoming)
        os.replace(incoming, path)
    finally:
        incoming.unlink(missing_ok=True)     # a no-op once the rename succeeded


class LocalSink:
    """SheetSink that writes to a local .xlsx workbook (folder/workbook.xlsx)."""

    def __init__(self) -> None:
        # The workbook an open batch() holds in memory and the file it belongs
        # to. Both stay None outside a batch, where write_tab does its own
        # read/write and is therefore still a complete operation on its own.
        self._book: Workbook | None = None
        self._path: Path | None = None
        self._fresh = False
        self._dirty = False

    def ensure_workbook(self, folder: str, workbook: str) -> Path:
        path = Path(folder).expanduser() / f"{workbook}.xlsx"
        path.parent.mkdir(parents=True, exist_ok=True)
        return path

    @contextmanager
    def batch(self, path: Path) -> Iterator[None]:
        """Hold the workbook open across a run of write_tab calls: one read, one write.

        write_tab used to load_workbook() and save() the ENTIRE .xlsx per tab,
        so publishing a single source (prices, details, history, about) parsed
        and rewrote the whole file four times over, and every source already in
        the book made the next export slower still — quadratic work for a linear
        job. What an export consists of is known before it starts
        (publish.workbook_tables), so the file is read once and written once.

        Nothing is saved unless the block finishes: an export that dies halfway
        leaves the workbook on disk exactly as it was, instead of carrying half
        of a re-export.
        """
        path = Path(path)
        if self._book is not None:
            yield                    # re-entrant: the outermost batch owns the save
            return
        self._book, self._fresh = self._load(path)
        self._path, self._dirty = path, False
        try:
            yield
            if self._dirty:          # an empty batch must not create a workbook
                _save(self._book, path)
        finally:
            self._book, self._path = None, None
            self._fresh, self._dirty = False, False

    def write_tab(self, path: Path, tab: str, header: list[str], rows: list[list]) -> None:
        path = Path(path)
        if self._book is not None and path == self._path:
            _write_sheet(self._book, tab, header, rows, fresh=self._fresh)
            self._dirty = True
            return
        # No batch holding THIS workbook: read, write, save, as it always did.
        book, fresh = self._load(path)
        _write_sheet(book, tab, header, rows, fresh=fresh)
        _save(book, path)

    def _load(self, path: Path) -> tuple[Workbook, bool]:
        """(workbook, fresh) — an absent file means a new book, hence a default
        sheet that has to go; an existing one is opened so its other tabs live."""
        xl = _openpyxl()
        return (xl.load_workbook(path), False) if path.exists() else (xl.Workbook(), True)

    def location(self, path: Path) -> str:
        return str(path)
