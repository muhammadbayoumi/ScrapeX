"""Local .xlsx export — same data + arrangement as the Google sink, on disk."""
from __future__ import annotations

import sqlite3
from pathlib import Path

import pytest

pytest.importorskip("openpyxl")
from openpyxl import load_workbook  # noqa: E402

from scrapex import db as dbmod  # noqa: E402
from scrapex.ingest import ingest_payloads  # noqa: E402
from scrapex.localsheets import LocalSink, _safe_title  # noqa: E402
from scrapex.publish import publish_source, workbook_tables  # noqa: E402
from scrapex.reports import EXPORT_HEADER  # noqa: E402
from tests.test_ingest import make_entry, make_payload, one_row  # noqa: E402


@pytest.fixture()
def conn() -> sqlite3.Connection:
    c = dbmod.connect(":memory:")
    dbmod.migrate(c)
    ingest_payloads(c, make_entry(), [make_payload([
        one_row(external_product_id="1", external_variant_id="v1", product_name="LED 400W",
                price="1,200.00"),
        one_row(external_product_id="2", external_variant_id="v2", product_name="Copper Wire",
                price="50.00", availability="out_of_stock"),
    ])])
    yield c
    c.close()


def test_export_creates_workbook_with_source_tab(tmp_path: Path, conn):
    n, location = publish_source(conn, "ELSEWEDYSHOP", LocalSink(), str(tmp_path), "ScrapeX Data")
    assert n == 2
    path = Path(location)
    assert path.exists() and path.name == "ScrapeX Data.xlsx"

    wb = load_workbook(path)
    assert "ELSEWEDYSHOP" in wb.sheetnames
    assert "Sheet" not in wb.sheetnames  # default empty sheet removed
    ws = wb["ELSEWEDYSHOP"]
    assert [c.value for c in ws[1]] == EXPORT_HEADER          # same header as Google
    names = {ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)}
    assert names == {"LED 400W", "Copper Wire"}
    # numeric price stays numeric (not a string) — same as the Google sink
    price_col = EXPORT_HEADER.index("price") + 1
    assert ws.cell(row=2, column=price_col).value in (1200.0, 50.0)


def test_export_is_idempotent_replace(tmp_path: Path, conn):
    LocalSink().ensure_workbook  # noqa: B018 - smoke
    publish_source(conn, "ELSEWEDYSHOP", LocalSink(), str(tmp_path), "ScrapeX Data")
    publish_source(conn, "ELSEWEDYSHOP", LocalSink(), str(tmp_path), "ScrapeX Data")
    wb = load_workbook(tmp_path / "ScrapeX Data.xlsx")
    assert wb.sheetnames.count("ELSEWEDYSHOP") == 1  # replaced, not duplicated


def test_second_source_adds_a_tab(tmp_path: Path, conn):
    sink = LocalSink()
    publish_source(conn, "ELSEWEDYSHOP", sink, str(tmp_path), "ScrapeX Data")
    # a second (empty->skip): simulate another source by writing directly
    sink.write_tab(tmp_path / "ScrapeX Data.xlsx", "MASDAR", EXPORT_HEADER, [["x"] * len(EXPORT_HEADER)])
    wb = load_workbook(tmp_path / "ScrapeX Data.xlsx")
    # One workbook, a tab per source — plus that source's own history and
    # about tabs. (Its details tab is skipped: this fixture publishes no
    # attributes, and a header with no rows is furniture, not data. The about
    # tab is never skipped: a workbook that does not say where its numbers came
    # from is what the owner opened and could not read.)
    assert set(wb.sheetnames) == {"ELSEWEDYSHOP", "ELSEWEDYSHOP — history",
                                  "ELSEWEDYSHOP — about", "MASDAR"}


def _count_workbook_io(monkeypatch) -> dict[str, int]:
    """Count how many times the .xlsx is parsed and rewritten from here on.

    Patched on the openpyxl module itself, which is what localsheets reaches for
    at call time; this file's own `load_workbook` was bound at import and so
    never shows up in the tally.
    """
    import openpyxl

    counts = {"load": 0, "save": 0}
    real_load, real_save = openpyxl.load_workbook, openpyxl.Workbook.save

    def load(*args, **kwargs):
        counts["load"] += 1
        return real_load(*args, **kwargs)

    def save(self, *args, **kwargs):
        counts["save"] += 1
        return real_save(self, *args, **kwargs)

    monkeypatch.setattr(openpyxl, "load_workbook", load)
    monkeypatch.setattr(openpyxl.Workbook, "save", save)
    return counts


def test_an_export_of_three_tabs_reads_and_writes_the_file_once(tmp_path: Path, conn,
                                                                monkeypatch):
    """One export costs one read and one write of the workbook, whatever its tab count.

    write_tab used to load_workbook() and save() the ENTIRE .xlsx per tab, so
    this export re-parsed and rewrote the whole book once per table it carries —
    and every source already in the workbook made the next export slower.
    """
    sink = LocalSink()
    publish_source(conn, "ELSEWEDYSHOP", sink, str(tmp_path), "ScrapeX Data")
    assert len(workbook_tables(conn, "ELSEWEDYSHOP")) == 3   # prices, history, about

    counts = _count_workbook_io(monkeypatch)
    publish_source(conn, "ELSEWEDYSHOP", sink, str(tmp_path), "ScrapeX Data")
    assert counts == {"load": 1, "save": 1}


def test_re_exporting_one_source_leaves_every_other_tab_intact(tmp_path: Path, conn):
    """The README's contract: re-running replaces THAT source's tab, only that one."""
    sink = LocalSink()
    path = tmp_path / "ScrapeX Data.xlsx"
    publish_source(conn, "ELSEWEDYSHOP", sink, str(tmp_path), "ScrapeX Data")
    sink.write_tab(path, "MASDAR", EXPORT_HEADER, [["kept"] + ["x"] * (len(EXPORT_HEADER) - 1)])

    publish_source(conn, "ELSEWEDYSHOP", sink, str(tmp_path), "ScrapeX Data")

    wb = load_workbook(path)
    assert set(wb.sheetnames) == {"ELSEWEDYSHOP", "ELSEWEDYSHOP — history",
                                  "ELSEWEDYSHOP — about", "MASDAR"}
    assert wb["MASDAR"].cell(row=2, column=1).value == "kept"    # not merely present
    assert wb.sheetnames.count("ELSEWEDYSHOP") == 1              # replaced, not duplicated


def test_an_export_that_dies_midway_leaves_the_previous_workbook_untouched(tmp_path: Path,
                                                                          conn):
    """A batch saves once at the end, so an interrupted one must save nothing.

    The workbook the owner already has is the thing being overwritten; half a
    re-export in it would be worse than no re-export at all.
    """
    sink = LocalSink()
    publish_source(conn, "ELSEWEDYSHOP", sink, str(tmp_path), "ScrapeX Data")
    path = tmp_path / "ScrapeX Data.xlsx"
    before = path.read_bytes()

    with pytest.raises(RuntimeError, match="crawl died"):
        with sink.batch(path):
            sink.write_tab(path, "MASDAR", EXPORT_HEADER, [["x"] * len(EXPORT_HEADER)])
            raise RuntimeError("the crawl died mid-export")

    assert path.read_bytes() == before
    assert "MASDAR" not in load_workbook(path).sheetnames
    # The next export still works: the failed batch let go of the workbook.
    sink.write_tab(path, "MASDAR", EXPORT_HEADER, [["x"] * len(EXPORT_HEADER)])
    assert "MASDAR" in load_workbook(path).sheetnames


def test_safe_title_truncates_and_sanitizes():
    assert _safe_title("A" * 40) == "A" * 31
    assert "/" not in _safe_title("A/B:C")


def test_publish_empty_source_raises(tmp_path: Path):
    c = dbmod.connect(":memory:")
    try:
        dbmod.migrate(c)
        with pytest.raises(ValueError, match="nothing to publish"):
            publish_source(c, "ELSEWEDYSHOP", LocalSink(), str(tmp_path), "ScrapeX Data")
    finally:
        c.close()
