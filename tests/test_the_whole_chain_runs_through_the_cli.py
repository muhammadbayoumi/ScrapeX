"""THE FIRST END-TO-END TEST THIS REPOSITORY HAS EVER HAD.

`ENGINEERING.md`'s own test matrix names end-to-end and chaos coverage (T7) and
`REVIEW-2026-07-28` §9 recorded that neither exists. Nearly two thousand tests
drive the pieces; not one drives the product. This drives the product:

    scrapex crawl  ->  scrapex ingest  ->  scrapex export

through `python -m scrapex.cli`, in a subprocess, against a shop served over real
HTTP on this machine, and then opens the workbook and reads the price back. Every
one of those steps has its own unit tests. What none of them can see is whether
the four fit together, which is the only thing an owner ever experiences.

WHY THIS COULD NOT BE WRITTEN BEFORE. `crawl`, `ingest`, `export` and `peek` all
call `load_manifest()` with no argument, so they read the repository's own
`sources.yaml` and nothing else — twelve real shops, every one of them a live
website. `SCRAPEX_SOURCES` (see `scrapex/config.load_manifest`) is what lets a
test hand the CLI a shop it controls.

WHAT IT WOULD HAVE CAUGHT. OP-1: three crawl jobs died on a machine running an
engine whose `jobs.py` had a call and not yet its helpers, and the suite was
green throughout, because every unit under test was fine and nothing ran the
chain. That is the shape of fault this file exists for.
"""
from __future__ import annotations

import json
import os
import subprocess
import sys
import threading
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path

import pytest

ROOT = Path(__file__).resolve().parent.parent

# The shop, as its API publishes it. Two products so an aggregate cannot pass by
# accident, and prices that are not round so a default cannot impersonate one.
PRODUCTS = [
    {"product_id": 101, "product_enname": "Test Mortar 25kg",
     "product_arname": "مونة اختبار ٢٥ كجم", "price": 137.5, "stock": 4,
     "sku": "TM-101"},
    {"product_id": 102, "product_enname": "Test Sealant 600ml",
     "product_arname": "مادة لحام اختبار ٦٠٠ مل", "price": 89, "stock": 0,
     "sku": "TS-102"},
]
BY_ID = {str(p["product_id"]): p for p in PRODUCTS}


class _Shop(BaseHTTPRequestHandler):
    """The two endpoints `custom-json-api` actually calls.

    The list is enveloped and the detail is bare, which is not a whim: it is
    what sikaegshop does, recorded in the connector's own `_detail` docstring
    after it was verified live. A stub that answered both the same way would
    make the connector's shape-handling untested here.
    """

    def do_GET(self) -> None:                              # noqa: N802
        path = self.path.split("?", 1)[0]
        if path == "/api/products":
            body = {"data": PRODUCTS, "pagination": {"totalPages": 1}}
        elif path.startswith("/api/products/"):
            product = BY_ID.get(path.rsplit("/", 1)[-1])
            body = product if product else None
        else:
            self.send_error(404)
            return
        if body is None:
            self.send_error(404)
            return
        raw = json.dumps(body).encode("utf-8")
        self.send_response(200)
        self.send_header("Content-Type", "application/json")
        self.send_header("Content-Length", str(len(raw)))
        self.end_headers()
        self.wfile.write(raw)

    def log_message(self, *_args) -> None:                 # keep pytest readable
        pass


@pytest.fixture
def shop():
    server = ThreadingHTTPServer(("127.0.0.1", 0), _Shop)
    thread = threading.Thread(target=server.serve_forever, daemon=True)
    thread.start()
    try:
        yield f"http://127.0.0.1:{server.server_address[1]}"
    finally:
        server.shutdown()
        server.server_close()


@pytest.fixture
def manifest(tmp_path, shop) -> Path:
    path = tmp_path / "sources.yaml"
    path.write_text(f"""
sources:
  - source_key: TESTSHOP
    source_name: Test Shop
    base_url: {shop}
    family: custom-json-api
    cadence: daily
    authority: shop
    active: true
    currency: EGP
    default_region: EG
    vat_mode: excl
    extract:
      - kind: product_prices
        scope: census
""", encoding="utf-8")
    return path


def cli(*args: str, manifest: Path, cwd: Path) -> subprocess.CompletedProcess:
    """The CLI as the owner's machine runs it: a separate process, no imports.

    Driving `cli.main()` in-process would share this interpreter's already
    imported modules and its `sys.argv`, and would not notice an entry point
    that cannot start — which is exactly the failure `packaging/engine_entry.py`
    shipped twice.
    """
    env = dict(os.environ, SCRAPEX_SOURCES=str(manifest))
    return subprocess.run(
        [sys.executable, "-m", "scrapex.cli", *args],
        capture_output=True, text=True, encoding="utf-8", errors="replace",
        cwd=str(cwd), env=env, timeout=300)


def test_a_price_travels_from_the_shop_to_the_workbook(tmp_path, manifest):
    """One price, one journey, four commands, and it is read back at the end.

    The assertion is on the NUMBER the shop published, not on a row count or an
    exit code. A chain can exit 0 at every step and still deliver an empty
    workbook, and that is the failure worth catching.
    """
    inbox, db = tmp_path / "inbox", tmp_path / "engine.db"
    out = tmp_path / "out"
    out.mkdir()

    crawled = cli("crawl", "TESTSHOP", "--inbox", str(inbox),
                  manifest=manifest, cwd=ROOT)
    assert crawled.returncode == 0, f"crawl failed:\n{crawled.stdout}\n{crawled.stderr}"
    assert list(inbox.rglob("*")), "crawl reported success and wrote nothing"

    ingested = cli("ingest", "TESTSHOP", "--inbox", str(inbox), "--db", str(db),
                   manifest=manifest, cwd=ROOT)
    assert ingested.returncode == 0, f"ingest failed:\n{ingested.stdout}\n{ingested.stderr}"
    assert db.exists(), "ingest reported success and there is no database"

    exported = cli("export", "TESTSHOP", "--db", str(db), "--folder", str(out),
                   manifest=manifest, cwd=ROOT)
    assert exported.returncode == 0, f"export failed:\n{exported.stdout}\n{exported.stderr}"

    books = list(out.rglob("*.xlsx"))
    assert books, f"export reported success and wrote no workbook:\n{exported.stdout}"

    from openpyxl import load_workbook

    cells = {str(cell.value) for sheet in load_workbook(books[0]).worksheets
             for row in sheet.iter_rows() for cell in row if cell.value is not None}

    assert "Test Mortar 25kg" in cells, (
        "the product the shop published is not in the workbook the owner opens")
    assert any(value in cells for value in ("137.5", "137,5")), (
        f"the price never arrived. What did: {sorted(cells)[:40]}")


def test_the_chain_refuses_a_source_the_manifest_does_not_have(tmp_path, manifest):
    """The other half of end to end: it must fail LOUDLY, not quietly.

    A crawl for an unknown key that exits 0 and writes nothing is
    indistinguishable, to everything downstream, from a shop that published
    nothing that day — and that is how an empty warehouse gets called a
    successful run.
    """
    result = cli("crawl", "NOSUCHSHOP", "--inbox", str(tmp_path / "inbox"),
                 manifest=manifest, cwd=ROOT)

    assert result.returncode != 0, (
        "an unknown source exited 0 — a run that collected nothing reported "
        "success")
    assert "NOSUCHSHOP" in (result.stdout + result.stderr), (
        "the refusal does not name what it refused")
